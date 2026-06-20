"""Text extraction from PDF, XLSX, and text files (ported from legacy kb_service)."""

import io
import logging
import os

logger = logging.getLogger(__name__)

EXCEL_ROWS_PER_CHUNK = 100

_EXT_MIME: dict[str, str] = {
    ".pdf": "application/pdf",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
}


def _extract_image(file_content: bytes) -> str:
    """Best-effort OCR for raster images; returns empty text if OCR is unavailable."""
    try:
        import pytesseract
        from PIL import Image

        with Image.open(io.BytesIO(file_content)) as img:
            text = pytesseract.image_to_string(img)
            return text.strip()
    except Exception:
        logger.warning("Image OCR unavailable; storing KB document without extracted text.")
        return ""


def _resolve_mime(file_name: str, mime_type: str) -> str:
    """Prefer extension-derived MIME -- browsers send unreliable types for xlsx."""
    ext = os.path.splitext(file_name)[1].lower()
    ext_mime = _EXT_MIME.get(ext, "")
    if ext_mime:
        return ext_mime
    if mime_type and mime_type not in ("application/octet-stream", ""):
        return mime_type
    return mime_type


def _extract_excel(file_content: bytes) -> tuple[str, list[str]]:
    """
    Read an Excel file and return:
      - full_text: complete Markdown representation
      - row_chunks: list of Markdown snippets, each covering EXCEL_ROWS_PER_CHUNK rows
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

    full_parts: list[str] = []
    row_chunks: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(c) if c is not None else "" for c in rows[0]]
        header_sep = [" --- " for _ in header]
        header_md = "| " + " | ".join(header) + " |"
        sep_md = "| " + " | ".join(header_sep) + " |"

        data_rows = rows[1:]

        # Full markdown
        md_lines = [header_md, sep_md]
        for row in data_rows:
            cells = [str(c) if c is not None else "" for c in row]
            while len(cells) < len(header):
                cells.append("")
            md_lines.append("| " + " | ".join(cells[: len(header)]) + " |")
        full_parts.append(f"## Sheet: {sheet_name}\n\n" + "\n".join(md_lines))

        # Row-based chunks
        for start in range(0, len(data_rows), EXCEL_ROWS_PER_CHUNK):
            batch = data_rows[start : start + EXCEL_ROWS_PER_CHUNK]
            chunk_lines = [
                f"## Sheet: {sheet_name} (rows {start + 1}-{start + len(batch)})",
                header_md,
                sep_md,
            ]
            for row in batch:
                cells = [str(c) if c is not None else "" for c in row]
                while len(cells) < len(header):
                    cells.append("")
                chunk_lines.append("| " + " | ".join(cells[: len(header)]) + " |")
            row_chunks.append("\n".join(chunk_lines))

    wb.close()

    if not full_parts:
        raise ValueError("The Excel file contains no data.")

    return "\n\n".join(full_parts), row_chunks


def _extract_pdf(file_content: bytes) -> str:
    """Extract text from a PDF using pdfplumber."""
    import pdfplumber

    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        pages_text = [page.extract_text() for page in pdf.pages]

    return "\n\n".join(text for text in pages_text if text)


class TextExtractor:
    """Extracts text from various file types."""

    def extract(
        self,
        file_content: bytes,
        file_name: str,
        mime_type: str,
    ) -> tuple[str, list[str] | None]:
        """
        Extract text from file content.

        Returns:
            (full_text, pre_chunks) where pre_chunks is set only for Excel
            (row-based chunks); None means use the default character chunker.
        """
        effective_mime = _resolve_mime(file_name, mime_type)

        if "text" in effective_mime or effective_mime in ("text/plain", "text/markdown"):
            return file_content.decode("utf-8", errors="replace"), None

        if "pdf" in effective_mime:
            return _extract_pdf(file_content), None

        if effective_mime.startswith("image/"):
            return _extract_image(file_content), None

        if (
            "spreadsheet" in effective_mime
            or "excel" in effective_mime
            or effective_mime
            in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            )
        ):
            full_text, row_chunks = _extract_excel(file_content)
            return full_text, row_chunks

        from src.common.domain.exceptions.knowledge_base import KBUnsupportedFileTypeError

        raise KBUnsupportedFileTypeError(effective_mime)
